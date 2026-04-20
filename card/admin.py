from django.contrib import admin
from import_export.admin import ImportExportModelAdmin
from django.shortcuts import render, redirect
from django.urls import path
from django.contrib import messages
from decimal import Decimal, ROUND_DOWN
import openpyxl

from .models import Card
from .resource import CardResource
from .forms import ExcelImport
from .ai_logic import clean_data_with_ai
from .utils import card_mask, phone_mask, send_admin_notification, format_card, format_phone


@admin.register(Card)
class CardAdmin(ImportExportModelAdmin):
    resource_classes = (CardResource,)

    list_display = ['masked_card', 'masked_phone', 'balance', 'status', 'expire']
    list_filter = ['status', 'expire']
    search_fields = ['card_number', 'phone']

    change_list_template = "admin/card/card/change_list.html"

    @admin.display(description="Card Number")
    def masked_card(self, obj):
        return card_mask(obj.card_number)

    @admin.display(description="Phone")
    def masked_phone(self, obj):
        return phone_mask(obj.phone)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name='import_excel'),
            path('ai-import/',    self.admin_site.admin_view(self.ai_import_view), name='ai_import'),
        ]
        return custom_urls + urls

    # --- AI Import ---
    def ai_import_view(self, request):
        if request.method == "POST":
            file = request.FILES.get("ai_excel_file")
            if not file:
                self.message_user(request, "Файл не выбран!", level=messages.ERROR)
                return redirect("..")
            try:
                wb = openpyxl.load_workbook(file)
                sheet = wb.active

                # Первая строка — заголовок, пропускаем
                raw_data = []
                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    if i == 0:
                        continue  # пропускаем header
                    raw_data.append(str(row))

                print(f"📄 Строк для AI: {len(raw_data)}")
                clean_json_list = clean_data_with_ai("\n".join(raw_data))
                print(f"🤖 AI вернул: {clean_json_list}")

                if not clean_json_list or not isinstance(clean_json_list, list):
                    self.message_user(request, "AI не распознал формат данных.", level=messages.ERROR)
                    return redirect("..")

                created_count = 0
                updated_count = 0
                error_count = 0

                for item in clean_json_list:
                    print(f"➡️  Обрабатываю: {item}")

                    card = str(item.get('card_number', '')).strip()
                    expire = str(item.get('expire', '')).strip()

                    if not card or not expire:
                        print(f"⚠️  Пропускаю — нет card или expire: {item}")
                        continue

                    try:
                        clean_balance = Decimal(
                            str(item.get('balance', 0)).replace(',', '')
                        ).quantize(Decimal('0.01'), rounding=ROUND_DOWN)
                    except Exception:
                        clean_balance = Decimal('0.00')

                    raw_status = str(item.get('status', '')).lower().strip()
                    status = raw_status if raw_status in ['active', 'inactive', 'expired'] else 'inactive'

                    phone = str(item.get('phone', '') or '').strip()

                    try:
                        _, created = Card.objects.update_or_create(
                            card_number=card,
                            defaults=dict(
                                expire=expire,
                                phone=phone,
                                status=status,
                                balance=clean_balance,
                            )
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    except Exception as e:
                        error_count += 1
                        print(f"❌ Ошибка сохранения {card}: {e}")

                # Telegram уведомление
                send_admin_notification(
                    created_count + updated_count,
                    method=f"✨ AI Import (новых: {created_count}, обновлено: {updated_count}, ошибок: {error_count})"
                )

                self.message_user(
                    request,
                    f"✨ AI импорт завершён! Создано: {created_count}, обновлено: {updated_count}, ошибок: {error_count}"
                )

            except Exception as e:
                self.message_user(request, f"Критическая ошибка: {e}", level=messages.ERROR)
                print(f"💥 Критическая ошибка в AI import: {e}")

            return redirect("..")

        context = dict(self.admin_site.each_context(request), title="AI Excel Import")
        return render(request, "admin/ai_upload_form.html", context)

    # --- Excel Import ---
    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelImport(request.POST, request.FILES)
            if form.is_valid():
                file = request.FILES["excel_file"]
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                errors = []
                created_count = 0

                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    # Защита от пустых строк
                    if not any(row):
                        continue

                    card_number = row[0]
                    phone = row[1] if len(row) > 1 else None

                    try:
                        clean_card = format_card(str(card_number))
                        clean_phone = format_phone(str(phone)) if phone else ""

                        if len(clean_card) != 16:
                            raise ValueError(f"Неверная длина карты: {len(clean_card)}")

                        Card.objects.update_or_create(
                            card_number=clean_card,
                            defaults=dict(phone=clean_phone)
                        )
                        created_count += 1
                    except Exception as e:
                        errors.append(f"Строка {row_idx}: {e}")
                        print(f"⚠️  {e}")

                if errors:
                    messages.error(request, f"Ошибки в строках: {'; '.join(errors)}")

                messages.success(request, f"✅ Успешно импортировано: {created_count}")
                return redirect("..")
        else:
            form = ExcelImport()

        return render(request, "admin/excel_form.html", {"form": form})


